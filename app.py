import openpyxl


def chisquared(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet1']
    expected_value = 4.30347826

    header = sheet.cell(1, 7)
    header.value = '+Observed - Expected'

    header2 = sheet.cell(1, 8)
    header2.value = '+(Observed - Expected)^2'

    header3 = sheet.cell(1, 11)
    header3.value = '+Chi Squared Value'

    for row in range(2, 48):
        observed_value = sheet.cell(row, 2)
        observed_expected = observed_value.value - expected_value
        observed_expected_value = sheet.cell(row, 7)
        observed_expected_value.value = observed_expected

        squared = sheet.cell(row, 8)
        squared.value = observed_expected_value.value * observed_expected_value.value

        chi_squared = sheet.cell(row, 11)
        chi_squared.value = squared.value / expected_value

    wb.save(filename)


chisquared('stats chi squared backup.xlsx')